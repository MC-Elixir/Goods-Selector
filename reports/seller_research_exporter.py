"""
卖家研究清单导出
================

把 build_seller_shortlist 产出的卖家清单导出为：
    1. Excel  .xlsx —— 一行一个卖家，含用户关心的全部字段 + 适合理由
    2. JSON        —— 全量结构化数据，便于下游 / MCP 消费

输入 payload 结构（见 agent.seller_research_service.shortlist_payload）：
    {
      "niche_label": str, "keyword": str, "marketplace": str,
      "ruleset_version": str, "quality_summary": dict,
      "items": [public dict, ...], "excluded_items": [public dict, ...]
    }
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from loguru import logger

from config.settings import settings


# 用户明确要求的清单字段：卖家名称 / 代表产品 / 标题 / 品牌 / 价格 / 评分 /
# 评论数 / 上架时间 / 月销量 / 月销售额，以及适合类型与适合理由。
_HEADERS = [
    "排名", "卖家名称", "适合类型", "适合度评分",
    "代表产品ASIN", "代表产品标题", "品牌",
    "价格($)", "评分", "评论数", "上架时间", "上架月数",
    "月销量", "月销售额($)", "在售商品数", "商品数来源",
    "适合理由",
]
_EXCLUDED_HEADERS = [
    "卖家名称", "代表产品ASIN", "代表产品标题", "品牌",
    "价格($)", "评分", "评论数", "月销量", "月销售额($)", "排除原因",
]


def export_seller_research(
    payload: dict[str, Any],
    *,
    output_dir: Optional[Path] = None,
    basename: Optional[str] = None,
) -> dict[str, Path]:
    """Write both the Excel and JSON deliverables; return their paths."""
    output_dir = output_dir or settings.export_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = _slug(payload.get("niche_label") or payload.get("keyword") or "sellers")
    stem = basename or f"seller_research_{slug}_{stamp}"

    xlsx_path = export_seller_research_excel(payload, output_dir / f"{stem}.xlsx")
    json_path = export_seller_research_json(payload, output_dir / f"{stem}.json")
    logger.info(f"卖家研究清单已导出：{xlsx_path.name} / {json_path.name}")
    return {"xlsx": xlsx_path, "json": json_path}


def export_seller_research_excel(payload: dict[str, Any], output_path: Path) -> Path:
    """导出卖家清单为 Excel（合格卖家 + 排除卖家两个 sheet）。"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - environment guard
        raise ImportError("请安装 openpyxl：pip install openpyxl") from exc

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "适合卖家清单"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, header in enumerate(_HEADERS, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, item in enumerate(payload.get("items") or [], 2):
        for col, value in enumerate(_eligible_row(row_idx - 1, item), 1):
            cell = sheet.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=col in (6, len(_HEADERS)))

    _autofit(sheet, get_column_letter, _HEADERS)

    excluded = payload.get("excluded_items") or []
    if excluded:
        drop_sheet = workbook.create_sheet("排除卖家")
        for col, header in enumerate(_EXCLUDED_HEADERS, 1):
            cell = drop_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_idx, item in enumerate(excluded, 2):
            for col, value in enumerate(_excluded_row(item), 1):
                drop_sheet.cell(row=row_idx, column=col, value=value)
        _autofit(drop_sheet, get_column_letter, _EXCLUDED_HEADERS)

    workbook.save(output_path)
    return output_path


def export_seller_research_json(payload: dict[str, Any], output_path: Path) -> Path:
    """导出全量结构化 JSON。"""
    document = {
        "niche_label": payload.get("niche_label"),
        "keyword": payload.get("keyword"),
        "marketplace": payload.get("marketplace"),
        "ruleset_version": payload.get("ruleset_version"),
        "generated_at": datetime.utcnow().isoformat(),
        "quality_summary": payload.get("quality_summary") or {},
        "items": payload.get("items") or [],
        "excluded_items": payload.get("excluded_items") or [],
    }
    output_path.write_text(
        json.dumps(document, ensure_ascii=False, indent=2, default=_json_default) + "\n",
        encoding="utf-8",
    )
    return output_path


def _eligible_row(rank: int, item: dict[str, Any]) -> list[Any]:
    return [
        rank,
        item.get("seller") or "-",
        item.get("fit_category_label") or item.get("fit_category") or "-",
        item.get("fit_score"),
        item.get("representative_asin") or "-",
        item.get("representative_title") or "-",
        item.get("brand") or "-",
        _num(item.get("price")),
        _num(item.get("rating")),
        item.get("review_count"),
        item.get("launch_date") or "-",
        _num(item.get("launch_months")),
        item.get("monthly_sales"),
        _num(item.get("monthly_revenue")),
        item.get("seller_product_count"),
        "导出" if item.get("product_count_source") == "reported" else "样本估计",
        _reason_text(item),
    ]


def _excluded_row(item: dict[str, Any]) -> list[Any]:
    return [
        item.get("seller") or "-",
        item.get("representative_asin") or "-",
        item.get("representative_title") or "-",
        item.get("brand") or "-",
        _num(item.get("price")),
        _num(item.get("rating")),
        item.get("review_count"),
        item.get("monthly_sales"),
        _num(item.get("monthly_revenue")),
        "；".join(item.get("exclusion_reasons") or []) or "-",
    ]


def _reason_text(item: dict[str, Any]) -> str:
    ai_reason = (item.get("ai_reason") or "").strip()
    if ai_reason:
        return ai_reason
    reasons = item.get("fit_reasons") or []
    return "；".join(str(reason) for reason in reasons) if reasons else "-"


def _autofit(sheet: Any, get_column_letter: Any, headers: list[str]) -> None:
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        longest = len(str(headers[col - 1]))
        for cell in sheet[letter]:
            value = cell.value
            if value is not None:
                longest = max(longest, min(60, len(str(value))))
        sheet.column_dimensions[letter].width = min(60, max(8, longest + 2))


def _num(value: Any) -> Any:
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return value


def _slug(value: Any) -> str:
    text = re.sub(r"[^0-9a-zA-Z]+", "_", str(value or "").strip().casefold()).strip("_")
    return (text or "sellers")[:40]


def _json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")
