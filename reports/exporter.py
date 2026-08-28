"""
报告导出
========

正式交付为一个 Excel 工作簿；JSON 仅供后台消费。Markdown 导出函数保留
给诊断调用方，但不再由正式可恢复流水线自动发布。

candidates 为 PipelineRecord 列表（见 pipeline/orchestrator.py）。
"""
from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from loguru import logger

from config.settings import settings


def _supplier_raw_score(supplier, key: str):
    raw = getattr(supplier, "raw_data", None) if supplier else None
    value = (raw or {}).get(key)
    return round(float(value), 3) if value is not None else None


def _display_score(value) -> str:
    return "-" if value is None else f"{value:.3f}"


def _raw_sourcing_payload(record) -> dict[str, Any]:
    product = getattr(record, "product", None)
    raw = getattr(product, "raw_data", None)
    payload = (raw or {}).get("sourcing_evidence") if isinstance(raw, dict) else None
    return payload if isinstance(payload, dict) else {}


def _record_rejection_reasons(record) -> list[str]:
    reasons = getattr(record, "rejection_reasons", None)
    if not reasons:
        score = getattr(record, "score", None)
        reasons = getattr(score, "rejection_reasons", None) or []
    recommendation = getattr(getattr(record, "sourcing_slice", None), "recommendation", None)
    raw_recommendation = _raw_sourcing_payload(record).get("recommendation") or {}
    evidence_reasons = (
        getattr(recommendation, "rejection_reasons", None)
        or raw_recommendation.get("rejection_reasons")
        or []
    )
    return list(dict.fromkeys([*reasons, *evidence_reasons]))


def _record_review_status(record) -> str:
    reasons = _record_rejection_reasons(record)
    score = getattr(record, "score", None)
    if score is None and reasons:
        return "insufficient_evidence"
    if score is not None and getattr(score, "passed_hard_filter", False):
        return "passed"
    return "rejected"


def _evidence_payload(record) -> dict[str, Any]:
    slice_result = getattr(record, "sourcing_slice", None)
    raw_payload = _raw_sourcing_payload(record)
    if not slice_result and raw_payload:
        recommendation = raw_payload.get("recommendation") or {}
        return {
            "schema_version": raw_payload.get("schema_version") or "target-sourcing-evidence-v1",
            "run_ref": raw_payload.get("run_ref"),
            "query_plan_and_hit_rates": raw_payload.get("query_attempts") or [],
            "match_evidence": raw_payload.get("evaluated_matches") or [],
            "recommendation_status": recommendation.get("status"),
            "recommendation_reasons": recommendation.get("recommendation_reasons") or [],
            "evidence_rejection_reasons": recommendation.get("rejection_reasons") or [],
            "manual_verification_tasks": recommendation.get("manual_verification_tasks") or [],
        }
    recommendation = getattr(slice_result, "recommendation", None)
    status = getattr(recommendation, "status", None)
    status = getattr(status, "value", status)

    def _model(value):
        if hasattr(value, "model_dump"):
            return value.model_dump(mode="json")
        if is_dataclass(value):
            return asdict(value)
        return value

    matches = []
    if slice_result:
        evaluated = getattr(slice_result, "evaluated_matches", None)
        values = evaluated if evaluated is not None else [
            *getattr(slice_result, "accepted_matches", []),
            *getattr(slice_result, "rejected_matches", []),
        ]
        matches = [_model(item) for item in values]
    return {
        "schema_version": "2.0",
        "run_ref": getattr(slice_result, "run_ref", None),
        "query_plan_and_hit_rates": getattr(slice_result, "query_attempts", []),
        "match_evidence": matches,
        "recommendation_status": status,
        "recommendation_reasons": list(getattr(recommendation, "recommendation_reasons", []) or []),
        "evidence_rejection_reasons": list(getattr(recommendation, "rejection_reasons", []) or []),
        "manual_verification_tasks": list(getattr(recommendation, "manual_verification_tasks", []) or []),
    }


def _json_cell(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _supplier_raw_value(supplier, key: str):
    raw = getattr(supplier, "raw_data", None) if supplier else None
    return (raw or {}).get(key) if isinstance(raw, dict) else None


def _supplier_match_decision(supplier) -> tuple[str, str]:
    """Return an evidence-aware, human-readable supplier row decision."""
    if _supplier_invalid_for_decision(supplier):
        return "不可用", "缺少可验证的真实 1688 货源身份"

    raw = getattr(supplier, "raw_data", None) or {}
    spec = raw.get("spec_match") if isinstance(raw.get("spec_match"), dict) else {}
    conflicts = list(spec.get("conflicts") or [])
    method = str(getattr(supplier, "match_verification_method", "") or "").lower()
    if method == "heuristic_rejected" or conflicts:
        reason = "匹配冲突: " + ", ".join(conflicts) if conflicts else "语义匹配未通过"
        return "淘汰", reason

    match_score = getattr(supplier, "match_quality_score", None)
    profit_margin = raw.get("supplier_profit_margin")
    rank_score = raw.get("supplier_rank_score")
    missing = []
    if match_score is None:
        missing.append("匹配度")
    if getattr(supplier, "base_price_cny", None) is None:
        missing.append("采购价")
    if profit_margin is None:
        missing.append("利润")
    if missing:
        return "待核验", "缺少" + "/".join(missing) + "证据"

    if float(match_score) >= 0.55 and float(profit_margin) >= 0.20 and float(rank_score or 0) >= 0.62:
        return "推荐", "匹配度、利润率和综合排名分均达标"
    if float(match_score) >= 0.40 and float(profit_margin) >= 0.10:
        return "观察", "可继续核对规格、样品与物流成本"
    return "淘汰", "匹配度或预估利润偏低"


def _write_supplier_match_sheet(workbook, candidates: list) -> int:
    """Write one sortable row per Amazon-product/1688-supplier pair."""
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet("Amazon×1688完整匹配")
    headers = [
        "建议", "建议依据", "Amazon排名", "ASIN", "Amazon标题", "Amazon链接",
        "Amazon售价($)", "Amazon评分", "Amazon评论数", "BSR排名", "产品综合分",
        "需求得分", "竞争得分", "供应得分", "物流得分", "风险得分",
        "1688供应商排名", "1688货源标题", "1688供应商", "是否工厂", "货源来源",
        "1688货源链接", "1688采购价(CNY)", "MOQ", "1688月销", "回头率",
        "匹配度", "规格匹配度", "图片相似度", "候选质量分", "供应商质量分",
        "业务条件分", "利润得分", "匹配+利润综合分", "预估利润率", "预估净利润($)",
        "采购成本($)", "头程($)", "FBA费($)", "佣金($)", "广告费($)", "退货损耗($)",
        "汇率损耗($)", "总成本($)", "匹配方式", "已匹配规格", "缺失证据", "冲突项",
    ]
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    decision_fills = {
        "推荐": PatternFill("solid", fgColor="C6EFCE"),
        "观察": PatternFill("solid", fgColor="FFEB9C"),
        "待核验": PatternFill("solid", fgColor="DDEBF7"),
        "淘汰": PatternFill("solid", fgColor="FFC7CE"),
        "不可用": PatternFill("solid", fgColor="D9D9D9"),
    }
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_index = 2
    for amazon_rank, record in enumerate(candidates, 1):
        product = record.product
        score = getattr(record, "score", None)
        for supplier_rank, supplier in enumerate(getattr(record, "suppliers", None) or [], 1):
            raw = getattr(supplier, "raw_data", None) or {}
            spec = raw.get("spec_match") if isinstance(raw.get("spec_match"), dict) else {}
            decision, decision_reason = _supplier_match_decision(supplier)
            values = [
                decision, decision_reason, amazon_rank, getattr(product, "asin", None),
                getattr(product, "title", None), getattr(product, "listing_url", None),
                getattr(product, "price", None), getattr(product, "rating", None),
                getattr(product, "review_count", None), getattr(product, "bsr_rank", None),
                getattr(score, "total_score", None) if score else None,
                getattr(score, "demand_score", None) if score else None,
                getattr(score, "competition_score", None) if score else None,
                getattr(score, "supply_score", None) if score else None,
                getattr(score, "logistics_score", None) if score else None,
                getattr(score, "risk_score", None) if score else None,
                supplier_rank, getattr(supplier, "title_cn", None), getattr(supplier, "supplier_name", None),
                getattr(supplier, "is_factory", None), _supplier_source(supplier), getattr(supplier, "offer_url", None),
                getattr(supplier, "base_price_cny", None), getattr(supplier, "moq", None),
                getattr(supplier, "monthly_sales", None), getattr(supplier, "repeat_buyer_rate", None),
                getattr(supplier, "match_quality_score", None), spec.get("score"),
                getattr(supplier, "image_similarity", None), raw.get("supplier_candidate_score"),
                raw.get("supplier_quality_score"), raw.get("supplier_business_score"),
                raw.get("supplier_profit_score"), raw.get("supplier_rank_score"),
                raw.get("supplier_profit_margin"), raw.get("supplier_net_profit"),
                raw.get("supplier_purchase_cost"), raw.get("supplier_shipping_cost"),
                raw.get("supplier_fba_fee"), raw.get("supplier_commission"), raw.get("supplier_ad_cost"),
                raw.get("supplier_return_loss"), raw.get("supplier_exchange_loss"), raw.get("supplier_total_cost"),
                getattr(supplier, "match_verification_method", None),
                ", ".join(spec.get("matched") or []), ", ".join(spec.get("missing") or []),
                ", ".join(spec.get("conflicts") or []),
            ]
            for column, value in enumerate(values, 1):
                cell = ws.cell(row=row_index, column=column, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=column in {2, 5, 18, 46, 47, 48})
            ws.cell(row=row_index, column=1).fill = decision_fills[decision]
            ws.cell(row=row_index, column=1).font = Font(bold=True)
            for column in (6, 22):
                link = ws.cell(row=row_index, column=column)
                if isinstance(link.value, str) and link.value.startswith(("http://", "https://")):
                    link.hyperlink = link.value
                    link.style = "Hyperlink"
            for column in (26, 27, 28, 29, 30, 31, 32, 33, 34, 35):
                ws.cell(row=row_index, column=column).number_format = "0.0%"
            for column in (7, 23, 36, 37, 38, 39, 40, 41, 42, 43, 44):
                ws.cell(row=row_index, column=column).number_format = "0.00"
            row_index += 1

    widths = [10, 34, 10, 14, 42, 35, 12, 10, 12, 10, 12, 10, 10, 10, 10, 10,
              14, 42, 24, 10, 18, 38, 16, 10, 12, 10, 10, 12, 12, 12, 14, 12,
              12, 16, 14, 14, 14, 11, 11, 11, 11, 13, 13, 13, 18, 28, 28, 28]
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    last_row = row_index - 1
    if last_row >= 2:
        for column in (11, 27, 28, 30, 31, 32, 33, 34, 35):
            letter = get_column_letter(column)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{last_row}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B",
                ),
            )
    return max(last_row - 1, 0)


# ============================================================
# Excel
# ============================================================

def export_excel(candidates: list, output_path: Optional[Path] = None) -> Path:
    """导出唯一正式工作簿，覆盖发现、市场、匹配和待核验数据。"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请安装 openpyxl：pip install openpyxl")

    output_path = output_path or (
        settings.export_dir / f"candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amazon商品"

    # --- 表头 ---
    headers = [
        "ASIN", "标题", "品牌", "类目",
        "售价($)", "BSR排名", "评分", "评论数",
        "综合评分", "利润率", "净利润($)",
        "利润得分", "需求得分", "竞争得分", "供应得分", "物流得分", "风险得分",
        "采购成本($)", "头程($)", "FBA费($)", "佣金($)", "广告费($)", "退货损耗($)",
        "供应商数", "Top1供应商", "Top1货源链接", "Top1采购价(CNY)", "Top1 MOQ",
        "Top1来源", "Top1视觉相似", "Top1匹配分", "Top1候选分", "Top1供应商质量分", "Top1业务条件分",
        "通过筛选", "审核状态", "拒绝原因",
        "Schema版本", "Run Ref", "查询计划与命中率", "匹配证据", "推荐状态",
        "推荐原因", "证据拒绝原因", "人工核验任务",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- 数据行 ---
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill = PatternFill("solid", fgColor="FFE7E7")

    for row_idx, rec in enumerate(candidates, 2):
        p = rec.product
        pb = rec.profit
        sc = rec.score
        sups = rec.suppliers or []
        top_sup = sups[0] if sups else None

        passed = sc.passed_hard_filter if sc else False
        row_fill = green_fill if passed else red_fill

        def _v(obj, attr, default=""):
            v = getattr(obj, attr, default)
            return v if v is not None else default

        def _score(obj, attr):
            value = getattr(obj, attr, None) if obj else None
            return round(float(value), 3) if value is not None else ""

        def _raw_score(obj, key):
            value = _supplier_raw_score(obj, key)
            return value if value is not None else ""

        row_data = [
            _v(p, "asin"),
            _v(p, "title"),
            _v(p, "brand"),
            _v(p, "category"),
            _v(p, "price"),
            _v(p, "bsr_rank"),
            _v(p, "rating"),
            _v(p, "review_count"),
            round(sc.total_score, 1) if sc else "",
            f"{pb.profit_margin:.1%}" if pb else "",
            round(pb.net_profit, 2) if pb else "",
            round(sc.profit_score, 3) if sc else "",
            round(sc.demand_score, 3) if sc else "",
            round(sc.competition_score, 3) if sc else "",
            round(sc.supply_score, 3) if sc else "",
            round(sc.logistics_score, 3) if sc else "",
            round(sc.risk_score, 3) if sc else "",
            round(pb.purchase_cost, 2) if pb else "",
            round(pb.shipping_cost, 2) if pb else "",
            round(pb.fba_fee, 2) if pb else "",
            round(pb.commission, 2) if pb else "",
            round(pb.ad_cost, 2) if pb else "",
            round(pb.return_loss, 2) if pb else "",
            len(sups),
            _v(top_sup, "supplier_name") if top_sup else "",
            _v(top_sup, "offer_url") if top_sup else "",
            _v(top_sup, "base_price_cny") if top_sup else "",
            _v(top_sup, "moq") if top_sup else "",
            _supplier_source(top_sup),
            _score(top_sup, "image_similarity"),
            _score(top_sup, "match_quality_score"),
            _raw_score(top_sup, "supplier_candidate_score"),
            _raw_score(top_sup, "supplier_quality_score"),
            _raw_score(top_sup, "supplier_business_score"),
            "✓" if passed else "✗",
            _record_review_status(rec),
            ", ".join(_record_rejection_reasons(rec)),
        ]
        evidence = _evidence_payload(rec)
        row_data.extend([
            evidence["schema_version"], evidence["run_ref"],
            _json_cell(evidence["query_plan_and_hit_rates"]),
            _json_cell(evidence["match_evidence"]), evidence["recommendation_status"],
            _json_cell(evidence["recommendation_reasons"]),
            _json_cell(evidence["evidence_rejection_reasons"]),
            _json_cell(evidence["manual_verification_tasks"]),
        ])

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # --- 列宽自适应 ---
    col_widths = [12, 40, 15, 20, 8, 10, 6, 8,
                  10, 8, 10,
                  10, 10, 10, 10, 10, 10,
                  10, 8, 8, 8, 8, 10,
                  8, 25, 40, 12, 8,
                  12, 10, 10, 10, 12, 12,
                  8, 24, 25]
    col_widths.extend([12, 18, 35, 35, 24, 30, 30, 30])
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    supplier_row_count = _write_supplier_match_sheet(wb, candidates)
    _write_market_sheet(wb, candidates)
    _write_review_sheet(wb, candidates)
    _write_summary_sheet(wb, candidates, supplier_row_count)
    wb.save(output_path)
    logger.info(
        f"Excel 导出完成：{output_path}"
        f"（{len(candidates)} 个 Amazon 产品，{supplier_row_count} 条 1688 匹配）"
    )
    return output_path


def _style_simple_sheet(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor="1F4E79")
    for index, header in enumerate(headers, 1):
        cell = ws.cell(1, index, header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(index)].width = min(max(len(header) * 2, 12), 42)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(workbook, records: list, supplier_rows: int) -> None:
    ws = workbook.create_sheet("运行摘要")
    passed = sum(_record_review_status(record) == "passed" for record in records)
    pending = sum(_record_review_status(record) != "passed" for record in records)
    values = [
        ("工作流", "Amazon crawler → SellerSprite market evidence → SellerSprite 1688 sourcing → evaluation → Excel"),
        ("生成时间", datetime.now().isoformat(timespec="seconds")),
        ("Amazon站点", "US"),
        ("Amazon商品数", len(records)),
        ("1688匹配行数", supplier_rows),
        ("通过商品数", passed),
        ("未通过及待核验数", pending),
        ("1688候选正式来源", "sellersprite_1688"),
        ("说明", "JSON为后台数据；本工作簿为默认人工交付物。"),
    ]
    _style_simple_sheet(ws, ["项目", "值"])
    for row_index, row in enumerate(values, 2):
        ws.cell(row_index, 1, row[0])
        ws.cell(row_index, 2, row[1])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100


def _write_market_sheet(workbook, records: list) -> None:
    ws = workbook.create_sheet("卖家精灵市场数据")
    headers = [
        "ASIN", "主关键词", "月搜索量", "月购买量", "购买率", "关键词难度",
        "机会分", "预估月销量", "竞品数", "Top10均价", "Top10评论数",
        "头部收入占比", "证据来源", "原始数据",
    ]
    _style_simple_sheet(ws, headers)
    for row_index, record in enumerate(records, 2):
        market = getattr(record, "market", None)
        raw = getattr(market, "raw_data", None) if market else None
        raw = raw if isinstance(raw, dict) else {}
        row = [
            record.product.asin,
            getattr(market, "main_keyword", None),
            getattr(market, "search_volume_monthly", None),
            getattr(market, "monthly_purchases", None),
            getattr(market, "purchase_rate", None),
            getattr(market, "keyword_difficulty", None),
            getattr(market, "opportunity_score", None),
            getattr(market, "est_monthly_sales", None),
            getattr(market, "competing_listings", None),
            getattr(market, "avg_price_top10", None),
            getattr(market, "avg_review_count_top10", None),
            getattr(market, "top10_revenue_share", None),
            raw.get("source_type") or raw.get("source") or "sellersprite_browser_extension",
            _json_cell(raw),
        ]
        for column, value in enumerate(row, 1):
            ws.cell(row_index, column, value)
    for column in (5, 6, 7, 12):
        for row_index in range(2, len(records) + 2):
            ws.cell(row_index, column).number_format = "0.0%"


def _write_review_sheet(workbook, records: list) -> None:
    ws = workbook.create_sheet("未通过及待核验")
    headers = ["ASIN", "Amazon标题", "状态", "原因", "供应商数", "人工核验任务", "Amazon链接"]
    _style_simple_sheet(ws, headers)
    row_index = 2
    for record in records:
        status = _record_review_status(record)
        if status == "passed":
            continue
        evidence = _evidence_payload(record)
        values = [
            record.product.asin,
            record.product.title,
            status,
            ", ".join(_record_rejection_reasons(record)) or "关键证据或硬筛选结果未满足",
            len(record.suppliers or []),
            _json_cell(evidence["manual_verification_tasks"]),
            getattr(record.product, "listing_url", None),
        ]
        for column, value in enumerate(values, 1):
            ws.cell(row_index, column, value)
        link = ws.cell(row_index, 7)
        if isinstance(link.value, str) and link.value.startswith(("http://", "https://")):
            link.hyperlink = link.value
            link.style = "Hyperlink"
        row_index += 1


# ============================================================
# Markdown
# ============================================================

_MD_TEMPLATE = """\
# {title}

> ASIN: `{asin}` | 类目: {category} | 综合评分: **{total_score}** | {passed_badge}

---

## 基本信息

| 字段 | 值 |
|---|---|
| 售价 | ${price} |
| BSR 排名 | {bsr_rank} |
| 评分 / 评论数 | {rating} / {review_count} |
| 品牌 | {brand} |
| Listing | {listing_url} |

{image_section}

---

## 评分明细

| 维度 | 得分（0-1） | 权重 |
|---|---|---|
| 利润率 | {profit_score} | 20% |
| 市场需求 | {demand_score} | 25% |
| 竞争烈度 | {competition_score} | 20% |
| 货源稳定性 | {supply_score} | 15% |
| 物流友好度 | {logistics_score} | 10% |
| 风险等级 | {risk_score} | 10% |
| **综合** | **{total_score}** | — |

{rejection_section}

---

## 利润明细

| 费项 | 金额(USD) |
|---|---|
| 售价 | ${selling_price:.2f} |
| 采购成本 | -${purchase_cost:.2f} |
| 头程物流 | -${shipping_cost:.2f} |
| FBA 费用 | -${fba_fee:.2f} |
| 平台佣金 | -${commission:.2f} |
| 广告费 | -${ad_cost:.2f} |
| 退货损耗 | -${return_loss:.2f} |
| 汇率损耗 | -${exchange_loss:.2f} |
| **净利润** | **${net_profit:.2f}** |
| **净利率** | **{profit_margin:.1%}** |

---

## 推荐货源（Top {sup_count}）

{suppliers_section}

---

*生成时间：{generated_at}*
"""


def export_markdown(candidates: list, output_dir: Optional[Path] = None) -> list[Path]:
    """每个候选产品导出一份 .md 报告。"""
    output_dir = output_dir or settings.export_dir / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    paths: list[Path] = []
    for rec in candidates:
        p = rec.product
        pb = rec.profit
        sc = rec.score
        sups = rec.suppliers or []

        asin = getattr(p, "asin", "unknown")
        img_url = getattr(p, "main_image_url", None)
        image_section = f"![{asin}]({img_url})" if img_url else ""

        rejection_section = f"> 审核状态：{_record_review_status(rec)}"
        rejection_reasons = _record_rejection_reasons(rec)
        if rejection_reasons:
            rejection_section = (
                f"> 审核状态：{_record_review_status(rec)}\n> ⚠️ 未通过硬性筛选：{', '.join(rejection_reasons)}"
            )

        # 货源表格
        sup_lines = [
            "| # | 来源 | 供应商 | 货源链接 | 采购价(CNY) | MOQ | 月销 | 回头率 | 匹配分 | 候选分 | 供应商质量分 |",
            "|---|---|---|---|---|---|---|---|---|---|---|",
        ]
        for i, s in enumerate(sups[:5], 1):
            sup_lines.append(
                f"| {i} | {_supplier_source(s)} "
                f"| {getattr(s,'supplier_name','-')} "
                f"| [{getattr(s,'alibaba_offer_id','-')}]({getattr(s,'offer_url','#')}) "
                f"| {getattr(s,'base_price_cny','-')} "
                f"| {getattr(s,'moq','-')} "
                f"| {getattr(s,'monthly_sales','-')} "
                f"| {getattr(s,'repeat_buyer_rate','-')} "
                f"| {_display_score(getattr(s,'match_quality_score', None))} "
                f"| {_display_score(_supplier_raw_score(s, 'supplier_candidate_score'))} "
                f"| {_display_score(_supplier_raw_score(s, 'supplier_quality_score'))} |"
            )

        content = _MD_TEMPLATE.format(
            title=getattr(p, "title", asin),
            asin=asin,
            category=getattr(p, "category", "-"),
            total_score=round(sc.total_score, 1) if sc else "-",
            passed_badge="✅ 通过筛选" if (sc and sc.passed_hard_filter) else "❌ 未通过筛选",
            price=getattr(p, "price", "-"),
            bsr_rank=getattr(p, "bsr_rank", "-"),
            rating=getattr(p, "rating", "-"),
            review_count=getattr(p, "review_count", "-"),
            brand=getattr(p, "brand", "-"),
            listing_url=getattr(p, "listing_url", "#"),
            image_section=image_section,
            profit_score=round(sc.profit_score, 3) if sc else "-",
            demand_score=round(sc.demand_score, 3) if sc else "-",
            competition_score=round(sc.competition_score, 3) if sc else "-",
            supply_score=round(sc.supply_score, 3) if sc else "-",
            logistics_score=round(sc.logistics_score, 3) if sc else "-",
            risk_score=round(sc.risk_score, 3) if sc else "-",
            rejection_section=rejection_section,
            selling_price=pb.selling_price if pb else 0,
            purchase_cost=pb.purchase_cost if pb else 0,
            shipping_cost=pb.shipping_cost if pb else 0,
            fba_fee=pb.fba_fee if pb else 0,
            commission=pb.commission if pb else 0,
            ad_cost=pb.ad_cost if pb else 0,
            return_loss=pb.return_loss if pb else 0,
            exchange_loss=pb.exchange_loss if pb else 0,
            net_profit=pb.net_profit if pb else 0,
            profit_margin=pb.profit_margin if pb else 0,
            sup_count=min(len(sups), 5),
            suppliers_section="\n".join(sup_lines),
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
        evidence = _evidence_payload(rec)
        evidence_section = (
            "\n## 证据链\n\n"
            f"- Schema version: `{evidence['schema_version']}`\n"
            f"- Run ref: `{evidence['run_ref'] or '-'}`\n"
            f"- Recommendation status: `{evidence['recommendation_status'] or '-'}`\n"
            f"- Query plan and hit rates: `{_json_cell(evidence['query_plan_and_hit_rates'])}`\n"
            f"- Match evidence: `{_json_cell(evidence['match_evidence'])}`\n"
            f"- Recommendation reasons: `{_json_cell(evidence['recommendation_reasons'])}`\n"
            f"- Rejection reasons: `{_json_cell(evidence['evidence_rejection_reasons'])}`\n"
            f"- Manual verification tasks: `{_json_cell(evidence['manual_verification_tasks'])}`\n\n"
        )
        content = content.replace("\n---\n\n*生成时间", f"{evidence_section}\n---\n\n*生成时间")

        out = output_dir / f"{asin}.md"
        out.write_text(content, encoding="utf-8")
        paths.append(out)

    logger.info(f"Markdown 导出完成：{len(paths)} 份报告 → {output_dir}")
    return paths


# ============================================================
# JSON
# ============================================================

def export_json(candidates: list, output_path: Optional[Path] = None) -> Path:
    """导出全量数据为 JSON，供下游程序消费。"""
    output_path = output_path or (
        settings.export_dir / f"candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    )

    def _serialize(rec) -> dict:
        p = rec.product
        pb = rec.profit
        sc = rec.score
        market = getattr(rec, "market", None)
        sups = rec.suppliers or []
        raw = getattr(p, "raw_data", None) if p else None
        raw = raw if isinstance(raw, dict) else {}

        payload = {
            "review_status": _record_review_status(rec),
            "rejection_reasons": _record_rejection_reasons(rec),
            "product": {
                k: getattr(p, k, None)
                for k in ("asin", "marketplace", "title", "brand", "category",
                          "price", "bsr_rank", "rating", "review_count",
                          "weight_kg", "length_cm", "width_cm", "height_cm",
                          "main_image_url", "listing_url")
            },
            "source_mode": raw.get("source_mode"),
            "source_query": raw.get("source_query") or raw.get("source_keyword") or raw.get("source_category"),
            "source_keyword": raw.get("source_keyword"),
            "keyword_normalized": raw.get("keyword_normalized"),
            "source_rank": raw.get("source_rank"),
            "source_warning": raw.get("keyword_warning"),
            "profit": {
                "selling_price": pb.selling_price,
                "purchase_cost": pb.purchase_cost,
                "shipping_cost": pb.shipping_cost,
                "fba_fee": pb.fba_fee,
                "commission": pb.commission,
                "ad_cost": pb.ad_cost,
                "return_loss": pb.return_loss,
                "exchange_loss": pb.exchange_loss,
                "net_profit": round(pb.net_profit, 4),
                "profit_margin": round(pb.profit_margin, 4),
            } if pb else None,
            "score": {
                "total_score": sc.total_score,
                "profit_score": sc.profit_score,
                "demand_score": sc.demand_score,
                "competition_score": sc.competition_score,
                "supply_score": sc.supply_score,
                "logistics_score": sc.logistics_score,
                "risk_score": sc.risk_score,
                "passed_hard_filter": sc.passed_hard_filter,
                "rejection_reasons": sc.rejection_reasons,
            } if sc else None,
            "market": _market_payload(market),
            "suppliers": [_supplier_payload(s) for s in sups],
        }
        payload.update(_evidence_payload(rec))
        return payload

    data = [_serialize(r) for r in candidates]
    output_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    logger.info(f"JSON 导出完成：{output_path}（{len(data)} 条）")
    return output_path


def _supplier_payload(supplier) -> dict[str, Any]:
    payload = {
        k: getattr(supplier, k, None)
        for k in ("alibaba_offer_id", "supplier_name", "offer_url",
                  "base_price_cny", "moq", "monthly_sales",
                  "repeat_buyer_rate", "is_factory", "delivery_days",
                  "title_cn", "offer_image_url", "image_similarity",
                  "match_quality_score",
                  "match_verification_method", "raw_data")
    }
    payload["supplier_quality_score"] = _supplier_raw_score(supplier, "supplier_quality_score")
    payload["supplier_business_score"] = _supplier_raw_score(supplier, "supplier_business_score")
    payload["candidate_score"] = _supplier_raw_score(supplier, "supplier_candidate_score")
    payload["rank_score"] = _supplier_raw_score(supplier, "supplier_rank_score")
    payload["profit_score"] = _supplier_raw_score(supplier, "supplier_profit_score")
    payload["profit_margin"] = _supplier_raw_value(supplier, "supplier_profit_margin")
    payload["net_profit"] = _supplier_raw_value(supplier, "supplier_net_profit")
    payload["purchase_cost_usd"] = _supplier_raw_value(supplier, "supplier_purchase_cost")
    payload["total_cost_usd"] = _supplier_raw_value(supplier, "supplier_total_cost")
    payload["selection_decision"] = _supplier_match_decision(supplier)[0]
    payload["sourcing_source"] = _supplier_source(supplier)
    payload["invalid_for_decision"] = _supplier_invalid_for_decision(supplier)
    return payload


def _supplier_source(supplier) -> str:
    if not supplier:
        return ""
    raw = getattr(supplier, "raw_data", None) or {}
    source = getattr(supplier, "sourcing_source", None) or raw.get("source")
    if source:
        return str(source)
    method = str(getattr(supplier, "match_verification_method", "") or "").lower()
    if method == "mock":
        return "mock"
    return "unknown"


def _supplier_invalid_for_decision(supplier) -> bool:
    if not supplier:
        return False
    raw = getattr(supplier, "raw_data", None) or {}
    source = _supplier_source(supplier).lower()
    method = str(getattr(supplier, "match_verification_method", "") or "").lower()
    name = str(getattr(supplier, "supplier_name", "") or "").lower()
    offer_id = str(getattr(supplier, "alibaba_offer_id", "") or "")
    return bool(
        raw.get("invalid_for_decision")
        or source == "mock"
        or method == "mock"
        or "mock" in name
        or (offer_id and not offer_id.isdigit())
    )


def _market_payload(market) -> dict[str, Any] | None:
    if not market:
        return None
    if is_dataclass(market):
        data = asdict(market)
    else:
        keys = (
            "asin", "marketplace", "brand", "seller_name", "title",
            "bsr", "bsr_category", "est_daily_sales", "est_monthly_sales",
            "price", "currency", "rating", "review_count", "available_date",
            "has_a_plus", "is_best_seller", "is_amazon_choice",
            "competing_listings", "avg_price_top10", "avg_review_count_top10",
            "top10_revenue_share", "main_keyword", "search_volume_monthly",
            "monthly_purchases", "purchase_rate", "keyword_difficulty",
            "opportunity_score", "seasonality", "raw_data",
        )
        data = {key: getattr(market, key, None) for key in keys}
    available = data.get("available_date")
    if isinstance(available, datetime):
        data["available_date"] = available.isoformat()
    return {key: value for key, value in data.items() if value not in (None, "", [], {})}
