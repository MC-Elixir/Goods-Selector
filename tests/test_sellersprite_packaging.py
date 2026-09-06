from __future__ import annotations

import json
from dataclasses import replace
from datetime import datetime, timedelta, timezone
from types import SimpleNamespace

import openpyxl
import pytest

from agent.sellersprite_models import SellerSpriteLocatorProfile
from agent.sellersprite_packaging import apply_packaging_evidence, parse_packaging_panel
from agent.tools.sellersprite_browser import PlaywrightSellerSpriteSession, SellerSpriteWorkflowError
from analyzers.profit_model import InsufficientCostEvidence, calc_shipping_cost, load_profit_params
from crawlers.amazon_bsr import ProductDTO
from pipeline.orchestrator import PipelineRecord
from reports.exporter import export_excel, export_json
from tests.test_sellersprite_browser import FakePage, valid_profile

ASIN = "B0BWMM644H"
URL = f"https://www.amazon.com/dp/{ASIN}"
# Sanitized physical labels observed on the real ASIN; no account or cookies.
PANEL = (
    f"ASIN:{ASIN}\n商品重量:11.5 pounds (5.22 kg)商品尺寸:106\"W x 92.5\"H"
    "包装重量:13.05 pounds (5.92 kg)包装尺寸:56.1 x 5.4 x 5.4 inches (大号大件)"
)


def product():
    return ProductDTO(asin=ASIN, marketplace="US", title="9ft umbrella", price=60.25)


def test_package_bundle_uses_packing_not_open_umbrella_dimensions():
    p = product()
    p.weight_kg = 5.216308
    payload = parse_packaging_panel(PANEL, asin=ASIN, source_ref=URL)
    apply_packaging_evidence(p, payload)
    assert p.weight_kg == pytest.approx(13.05 * 0.453592)
    assert (p.length_cm, p.width_cm, p.height_cm) == pytest.approx((142.494, 13.716, 13.716))
    assert p.raw_data["logistics_evidence"]["applied"] is True
    assert payload["fields"]["package_dimensions"]["source_ref"] == URL
    assert calc_shipping_cost(p, load_profit_params()) > 0


@pytest.mark.parametrize("text", [
    f"ASIN:{ASIN} 商品尺寸:106 x 106 x 92.5 inches 商品重量:11.5 pounds",
    f"ASIN:{ASIN} 包装尺寸:56.1 x 5.4 x 5.4 inches",
    f"ASIN:{ASIN} 包装尺寸:106\"W x 92.5\"H 包装重量:13 pounds",
    f"ASIN:{ASIN} 包装尺寸:0 x 5 x 5 inches 包装重量:13 pounds",
])
def test_missing_package_evidence_cannot_use_item_values_or_impute(text):
    p = product()
    p.weight_kg, p.length_cm, p.width_cm, p.height_cm = (5, 270, 270, 235)
    apply_packaging_evidence(p, parse_packaging_panel(text, asin=ASIN, source_ref=URL))
    assert p.raw_data["logistics_evidence"]["applied"] is False
    with pytest.raises(InsufficientCostEvidence):
        calc_shipping_cost(p, load_profit_params())


def test_packaging_rejects_cross_asin_or_stale_evidence():
    with pytest.raises(ValueError, match="ASIN mismatch"):
        parse_packaging_panel(PANEL, asin=ASIN, source_ref="https://www.amazon.com/dp/B0CSFK27JW")
    assert parse_packaging_panel(PANEL.replace(ASIN, "B0CSFK27JW"), asin=ASIN, source_ref=URL) == {}
    payload = parse_packaging_panel(PANEL, asin=ASIN, source_ref=URL)
    payload["fields"]["package_dimensions"]["expires_at"] = (
        datetime.now(timezone.utc) - timedelta(seconds=1)
    ).isoformat()
    p = product()
    apply_packaging_evidence(p, payload)
    assert p.length_cm is None


def test_optional_profile_and_browser_read_are_same_asin_and_read_only(tmp_path, monkeypatch):
    profile = replace(valid_profile(), product_packaging="css=packaging")
    path = tmp_path / "profile.json"
    from dataclasses import asdict

    path.write_text(json.dumps(asdict(profile)), encoding="utf-8")
    assert SellerSpriteLocatorProfile.from_json(path).product_packaging == "css=packaging"
    page = FakePage(asin=ASIN, visible_markers={"packaging"})
    session = PlaywrightSellerSpriteSession(profile=profile, download_dir=tmp_path, page=page)
    monkeypatch.setattr(session, "_wait_until_visible", lambda *a, **kw: True)
    monkeypatch.setattr(session, "_target_locator", lambda *a: SimpleNamespace(
        is_visible=lambda: False, inner_text=lambda **kw: PANEL,
    ))
    payload = session.read_product_packaging(ASIN)
    assert payload["fields"]["package_dimensions"]["value"] == pytest.approx([142.494, 13.716, 13.716])
    assert page.clicked == []
    with pytest.raises(SellerSpriteWorkflowError, match="ASIN_MISMATCH"):
        session.read_product_packaging("B0CSFK27JW")


def test_export_preserves_package_evidence_and_explains_all_rejected(tmp_path):
    p = product()
    apply_packaging_evidence(p, parse_packaging_panel(PANEL, asin=ASIN, source_ref=URL))
    p.raw_data["rejected_suppliers"] = [{
        "alibaba_offer_id": "123", "supplier_name": "test supplier", "raw_data": {
            "spec_match": {"conflicts": ["material", "title_relevance"]},
        },
    }]
    rec = PipelineRecord(product=p)
    path = export_json([rec], tmp_path / "audit.json")
    result = json.loads(path.read_text(encoding="utf-8"))[0]
    assert result["review_status"] == "rejected"
    assert "no_qualified_suppliers" in result["rejection_reasons"]
    assert "supplier_spec_conflict:material" in result["rejection_reasons"]
    assert result["logistics_evidence"]["fields"]["package_dimensions"]["source_ref"] == URL
    wb = openpyxl.load_workbook(export_excel([rec], tmp_path / "audit.xlsx"), data_only=True)
    ws = wb["Amazon商品"]
    headers = [cell.value for cell in ws[1]]
    assert ws.cell(2, headers.index("包装长(cm)") + 1).value == pytest.approx(142.494)
    assert ws.cell(2, headers.index("包装证据来源") + 1).value == URL
    wb.close()


def test_formal_sourcing_carries_packaging_without_extra_query():
    from agent.sellersprite_1688_sourcing import run_sellersprite_1688_sourcing

    actions = []

    class Session:
        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def open_amazon_product(self, asin):
            actions.append(("open", asin))

        def check_sellersprite_extension(self):
            pass

        def read_product_packaging(self, asin):
            actions.append(("packaging", asin))
            return parse_packaging_panel(PANEL, asin=asin, source_ref=URL)

        def source_1688_suppliers(self, asin):
            actions.append(("source", asin))
            return []

    deps = SimpleNamespace(
        browser_enabled=True, profile=SimpleNamespace(has_sourcing_1688_locators=lambda: True),
        session_factory=Session, is_cancelled=lambda: False,
    )
    p = product()
    assert run_sellersprite_1688_sourcing(ASIN, product=p, dependencies=deps, required=True) == []
    assert p.raw_data["logistics_evidence"]["applied"]
    assert actions == [("open", ASIN), ("packaging", ASIN), ("source", ASIN)]


def test_resumed_profit_reuses_the_enriched_product_snapshot(monkeypatch):
    from config.settings import settings
    from db.models import ExecutionNode, ProfitSnapshot
    from matchers.alibaba_pailitao import SupplierDTO
    from pipeline.orchestrator import resume_pipeline, run_pipeline
    from tests.test_recoverable_pipeline import _memory_session_scope, _profit, _score

    Session, session_scope = _memory_session_scope()
    calls = {"match": 0, "profit": 0}

    def match(item, **kwargs):
        calls["match"] += 1
        apply_packaging_evidence(item, parse_packaging_panel(PANEL, asin=ASIN, source_ref=URL))
        return [SupplierDTO(alibaba_offer_id="package-test", base_price_cny=20, moq=10)]

    def profit(item, supplier):
        calls["profit"] += 1
        assert item.length_cm == pytest.approx(142.494)
        return _profit(item, supplier)

    monkeypatch.setattr("pipeline.orchestrator.session_scope", session_scope)
    monkeypatch.setattr("crawlers.amazon_bsr.crawl_best_sellers", lambda *a: [product()])
    monkeypatch.setattr("pipeline.recoverable._formal_match_suppliers", match)
    monkeypatch.setattr("pipeline.orchestrator.predict_profit", profit)
    monkeypatch.setattr("pipeline.orchestrator.score_product", _score)
    monkeypatch.setattr(settings, "mjjl_max_products_per_run", 0)
    run_id = run_pipeline("Home & Kitchen", limit=1, export=False)
    assert resume_pipeline(run_id) == run_id
    assert calls == {"match": 1, "profit": 1}
    with Session() as s:
        assert s.query(ProfitSnapshot).count() == 1
        node = s.query(ExecutionNode).filter_by(run_id=run_id, stage="profit").one()
        assert node.input_snapshot["product"]["length_cm"] == pytest.approx(142.494)
