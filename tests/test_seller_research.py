"""Tests for the market-research seller-shortlist workflow."""
from __future__ import annotations

from datetime import date
from types import SimpleNamespace

import openpyxl
import pytest
from sqlalchemy import create_engine

from agent.sellersprite_models import SellerSpriteLocatorProfile
from agent.tools.browser_downloads import DownloadedArtifact
from agent.tools.competitor_importer import (
    CompetitorImportError,
    import_competitor_export,
)
from analyzers.seller_research import (
    CompetitorRow,
    build_seller_shortlist,
    load_rules_config,
)
from db.migrate import run_migrations
from db.seller_research_repository import (
    get_seller_research_run,
    list_seller_research_runs,
    save_seller_research,
)

AS_OF = date(2026, 7, 1)


def _row(**kwargs) -> CompetitorRow:
    defaults = dict(
        seller="Seller",
        asin="B000000001",
        title="Item",
        brand="Brand",
        price=29.99,
        rating=4.2,
        review_count=150,
        launch_date="2025-09-01",
        monthly_sales=300,
        monthly_revenue=9000.0,
        seller_product_count=8,
    )
    defaults.update(kwargs)
    return CompetitorRow(**defaults)


def _write_csv(path, header, rows):
    lines = [",".join(header)]
    for row in rows:
        lines.append(",".join("" if value is None else str(value) for value in row))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


# ============================================================
# Rules config + engine
# ============================================================
def test_rules_config_loads_and_weights_sum_to_one():
    config = load_rules_config()
    weights = config["fit_score"]["weights"]
    assert abs(sum(weights.values()) - 1.0) < 1e-6
    assert config["version"]


def test_rules_config_rejects_bad_weights(tmp_path):
    bad = tmp_path / "bad.yaml"
    bad.write_text(
        "version: '0.0'\nexclusions: {}\ncategories: {}\n"
        "fit_score:\n  weights:\n    a: 0.5\n    b: 0.2\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError):
        load_rules_config(bad)


def test_classify_low_competition_efficient():
    rows = [_row(seller="Efficient", seller_product_count=8, monthly_revenue=12000, monthly_sales=400)]
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    item = shortlist.items[0]
    assert item.fit_category == "low_competition_efficient"
    assert item.fit_reasons and "少而精" in item.fit_reasons[0]
    assert 0 <= item.fit_score <= 100


def test_classify_new_rising():
    rows = [_row(seller="Fresh", seller_product_count=40, monthly_revenue=4000, monthly_sales=200, review_count=90, launch_date="2026-01-01")]
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    item = shortlist.items[0]
    assert item.fit_category == "new_rising"


def test_classify_differentiation_opportunity():
    rows = [_row(seller="Improve", seller_product_count=30, monthly_revenue=8000, monthly_sales=200, rating=4.0, review_count=800, launch_date="2020-01-01")]
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    item = shortlist.items[0]
    assert item.fit_category == "differentiation_opportunity"


def test_exclusion_head_seller_by_reviews():
    rows = [_row(seller="Giant", review_count=8000)]
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    assert not shortlist.items
    assert shortlist.excluded_items[0].excluded
    assert any("头部卖家" in reason for reason in shortlist.excluded_items[0].exclusion_reasons)


def test_exclusion_missing_required_field():
    rows = [_row(seller="Sparse", monthly_sales=None)]
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    assert not shortlist.items
    assert any("数据不足" in reason for reason in shortlist.excluded_items[0].exclusion_reasons)


def test_brand_monopoly_exclusion():
    rows = []
    # 4 of 10 rows share one brand across distinct sellers -> 0.4 share >= 0.35
    for index in range(4):
        rows.append(_row(seller=f"Mono{index}", brand="BigBrand", asin=f"B0000000{index}0"))
    for index in range(6):
        rows.append(_row(seller=f"Indie{index}", brand=f"Brand{index}", asin=f"B0000001{index}0"))
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    excluded_sellers = {item.seller for item in shortlist.excluded_items}
    assert {"Mono0", "Mono1", "Mono2", "Mono3"} <= excluded_sellers
    assert all(
        any("品牌垄断" in reason for reason in item.exclusion_reasons)
        for item in shortlist.excluded_items
        if item.seller.startswith("Mono")
    )


def test_fit_scores_stay_in_bounds():
    rows = [
        _row(seller="A", monthly_sales=5000, review_count=1, seller_product_count=1, launch_date="2026-06-01"),
        _row(seller="B", monthly_sales=90, review_count=1400, seller_product_count=55, launch_date="2020-01-01", rating=4.9),
    ]
    shortlist = build_seller_shortlist(rows, as_of=AS_OF)
    for item in shortlist.items:
        assert 0 <= item.fit_score <= 100


# ============================================================
# Importer
# ============================================================
def test_import_csv_english_headers(tmp_path):
    path = _write_csv(
        tmp_path / "en.csv",
        ["Seller", "ASIN", "Product Title", "Brand", "Price", "Rating", "Reviews", "Launch Date", "Monthly Sales", "Monthly Revenue", "Seller Products"],
        [["WoodCraft", "B0AAA11111", "Cedar House", "WoodCraft", "28.99", "4.3", "187", "2025-09-01", "420", "12176", "8"]],
    )
    imported = import_competitor_export(path, niche_label="birdhouse", keyword="bird house")
    assert imported.row_count == 1
    row = imported.competitor_rows[0]
    assert row.seller == "WoodCraft" and row.review_count == 187
    assert row.monthly_sales == 420 and row.monthly_revenue == 12176
    assert row.launch_date == "2025-09-01"


def test_import_csv_chinese_headers(tmp_path):
    path = _write_csv(
        tmp_path / "cn.csv",
        ["卖家", "ASIN", "商品标题", "品牌", "价格", "评分", "评论数", "上架时间", "月销量", "月销售额", "在售商品数"],
        [["巢艺工坊", "B0BBB22222", "雪松鸟屋", "巢艺", "35.99", "4.1", "92", "2026/01/15", "310", "11157", "5"]],
    )
    imported = import_competitor_export(path, niche_label="鸟屋", keyword="birdhouse")
    row = imported.competitor_rows[0]
    assert row.seller == "巢艺工坊" and row.brand == "巢艺"
    assert row.monthly_sales == 310 and row.launch_date == "2026-01-15"


def test_import_numeric_units_and_currency(tmp_path):
    path = _write_csv(
        tmp_path / "units.csv",
        ["Seller", "Price", "Reviews", "Monthly Sales", "Monthly Revenue"],
        [["Shop", "$1234.56", "3000", "1.2k", "12万"]],
    )
    imported = import_competitor_export(path)
    row = imported.competitor_rows[0]
    assert row.price == 1234.56
    assert row.review_count == 3000
    assert row.monthly_sales == 1200
    assert row.monthly_revenue == 120000


def test_import_seller_falls_back_to_brand(tmp_path):
    path = _write_csv(
        tmp_path / "brandonly.csv",
        ["Brand", "Price", "Monthly Sales"],
        [["OnlyBrand", "20", "150"]],
    )
    imported = import_competitor_export(path)
    assert imported.competitor_rows[0].seller == "OnlyBrand"


def test_import_missing_identity_raises(tmp_path):
    path = _write_csv(tmp_path / "noid.csv", ["Price", "Monthly Sales"], [["20", "150"]])
    with pytest.raises(CompetitorImportError) as exc:
        import_competitor_export(path)
    assert exc.value.error_code == "MISSING_IDENTITY"


def test_import_empty_rows_raises(tmp_path):
    path = _write_csv(tmp_path / "empty.csv", ["Seller", "Price"], [])
    with pytest.raises(CompetitorImportError) as exc:
        import_competitor_export(path)
    assert exc.value.error_code == "EMPTY_EXPORT"


def test_import_xlsx(tmp_path):
    path = tmp_path / "c.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Seller", "ASIN", "Price", "Rating", "Reviews", "Monthly Sales", "Monthly Revenue"])
    sheet.append(["XlShop", "B0CCC33333", 30.0, 4.2, 120, 250, 7500])
    workbook.save(path)
    imported = import_competitor_export(path, niche_label="x")
    assert imported.competitor_rows[0].seller == "XlShop"
    assert imported.competitor_rows[0].monthly_sales == 250


# ============================================================
# Service
# ============================================================
def test_run_from_import_without_ai(tmp_path):
    from agent.seller_research_service import run_seller_research_from_import

    path = _write_csv(
        tmp_path / "svc.csv",
        ["Seller", "Price", "Rating", "Reviews", "Launch Date", "Monthly Sales", "Monthly Revenue", "Seller Products"],
        [
            ["Small", "28.99", "4.2", "150", "2025-10-01", "400", "12000", "8"],
            ["Head", "59.99", "4.7", "9000", "2018-01-01", "2000", "119980", "200"],
        ],
    )
    imported = import_competitor_export(path, niche_label="birdhouse", keyword="bird house")
    payload = run_seller_research_from_import(imported, engine=None, generate_ai_reasons=False, export=False)
    assert [item["seller"] for item in payload["items"]] == ["Small"]
    assert payload["excluded_items"][0]["seller"] == "Head"
    assert payload["ai_reasons"]["status"] == "skipped"


def test_attach_ai_reasons_with_fake_client():
    from agent.seller_research_service import attach_ai_reasons

    shortlist = build_seller_shortlist([_row(seller="Alpha"), _row(seller="Beta", asin="B000000002")], as_of=AS_OF)

    class _FakeCompletions:
        def create(self, **_kwargs):
            content = '{"reasons": {"0": "低竞争值得研究", "1": "新品有机会"}}'
            return SimpleNamespace(choices=[SimpleNamespace(message=SimpleNamespace(content=content))])

    fake_client = SimpleNamespace(chat=SimpleNamespace(completions=_FakeCompletions()))
    status = attach_ai_reasons(shortlist.items, niche_label="birdhouse", client=fake_client)
    assert status["status"] == "success"
    assert shortlist.items[0].ai_reason == "低竞争值得研究"


def test_attach_ai_reasons_skips_without_key(monkeypatch):
    from agent import seller_research_service
    from config.settings import settings

    monkeypatch.setattr(settings, "ppio_api_key", "")
    shortlist = build_seller_shortlist([_row(seller="Alpha")], as_of=AS_OF)
    status = seller_research_service.attach_ai_reasons(shortlist.items)
    assert status["status"] == "skipped"
    assert shortlist.items[0].ai_reason is None


def test_parse_reasons_handles_reasoning_model_dual_output():
    from agent.seller_research_service import _parse_reasons

    # MiniMax-style reasoning models emit a bare JSON object followed by a
    # fenced copy; the first balanced object must win, not a span across both.
    content = '{"reasons":{"0":"少而精"}}\n\n```json\n{\n  "reasons": {\n    "0": "少而精"\n  }\n}\n```'
    assert _parse_reasons(content) == {"0": "少而精"}
    assert _parse_reasons("```json\n{\"reasons\": {\"1\": \"新品起量\"}}\n```") == {"1": "新品起量"}
    assert _parse_reasons("") == {}
    assert _parse_reasons("no json here") == {}


def test_run_competitor_export_extension_unavailable():
    from agent.seller_research_service import run_competitor_export

    deps = SimpleNamespace(browser_enabled=False, profile=None, session_factory=None)
    result = run_competitor_export("bird house", niche_label="birdhouse", dependencies=deps)
    assert result["status"] == "EXTENSION_UNAVAILABLE"


def test_run_competitor_export_happy_path(tmp_path):
    from agent.seller_research_service import run_competitor_export

    csv_path = _write_csv(
        tmp_path / "browser.csv",
        ["Seller", "Price", "Rating", "Reviews", "Launch Date", "Monthly Sales", "Monthly Revenue", "Seller Products"],
        [["BrowserShop", "30", "4.2", "150", "2025-11-01", "300", "9000", "6"]],
    )
    artifact = DownloadedArtifact.from_path(csv_path)

    class _FakeSession:
        def __init__(self):
            self.opened = []
            self.keyword = None

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def open_sellersprite_page(self, url):
            self.opened.append(url)

        def export_competitor_products(self, keyword):
            self.keyword = keyword
            return artifact

    session = _FakeSession()
    deps = SimpleNamespace(
        browser_enabled=True,
        profile=_competitor_profile(),
        session_factory=lambda: session,
    )
    result = run_competitor_export(
        "bird house",
        niche_label="birdhouse",
        sellersprite_url="https://www.sellersprite.com/x",
        dependencies=deps,
        engine=None,
        generate_ai_reasons=False,
        export=False,
    )
    assert result["status"] == "SUCCESS"
    assert result["items"][0]["seller"] == "BrowserShop"
    assert session.keyword == "bird house"
    assert session.opened == ["https://www.sellersprite.com/x"]


# ============================================================
# Repository + exporter
# ============================================================
def test_repository_save_get_list(tmp_path):
    from analyzers.seller_research import build_seller_shortlist as build

    engine = create_engine(f"sqlite:///{tmp_path / 'research.db'}")
    run_migrations(engine)
    path = _write_csv(
        tmp_path / "repo.csv",
        ["Seller", "Price", "Rating", "Reviews", "Launch Date", "Monthly Sales", "Monthly Revenue", "Seller Products"],
        [
            ["KeepMe", "28.99", "4.2", "150", "2025-10-01", "400", "12000", "8"],
            ["DropMe", "59.99", "4.7", "9000", "2018-01-01", "2000", "119980", "200"],
        ],
    )
    imported = import_competitor_export(path, niche_label="birdhouse", keyword="bird house")
    shortlist = build(imported.competitor_rows, as_of=AS_OF)

    saved = save_seller_research(engine, imported, shortlist, export_file="x.xlsx")
    assert saved["eligible_count"] == 1 and saved["excluded_count"] == 1
    assert saved["items"][0]["seller"] == "KeepMe"
    assert saved["excluded_items"][0]["seller"] == "DropMe"

    fetched = get_seller_research_run(engine, saved["id"])
    assert fetched["items"][0]["fit_reasons"]
    runs = list_seller_research_runs(engine, limit=10)
    assert runs and runs[0]["niche_label"] == "birdhouse"


def test_exporter_writes_excel_and_json(tmp_path):
    from agent.seller_research_service import run_seller_research_from_import
    from reports.seller_research_exporter import export_seller_research

    path = _write_csv(
        tmp_path / "exp.csv",
        ["Seller", "Price", "Rating", "Reviews", "Launch Date", "Monthly Sales", "Monthly Revenue", "Seller Products"],
        [["ExpShop", "28.99", "4.2", "150", "2025-10-01", "400", "12000", "8"]],
    )
    imported = import_competitor_export(path, niche_label="birdhouse", keyword="bird house")
    payload = run_seller_research_from_import(imported, engine=None, generate_ai_reasons=False, export=False)
    paths = export_seller_research(payload, output_dir=tmp_path)
    assert paths["xlsx"].exists() and paths["json"].exists()
    workbook = openpyxl.load_workbook(paths["xlsx"])
    sheet = workbook.active
    assert sheet.cell(row=1, column=2).value == "卖家名称"
    assert sheet.cell(row=2, column=2).value == "ExpShop"


# ============================================================
# Locator profile + browser session
# ============================================================
def _competitor_profile() -> SellerSpriteLocatorProfile:
    return SellerSpriteLocatorProfile(
        panel_open="css=panel_open",
        ready="css=ready",
        login_required="css=login_required",
        permission_required="css=permission_required",
        captcha="css=captcha",
        reverse_keywords="css=reverse_keywords",
        asin_input="css=asin_input",
        submit="css=submit",
        results_ready="css=results_ready",
        export_menu="css=export_menu",
        export="css=export",
        competitor_lookup="css=competitor_lookup",
        competitor_keyword_input="css=competitor_keyword_input",
        competitor_submit="css=competitor_submit",
        competitor_results_ready="css=competitor_results_ready",
        competitor_export_menu="css=competitor_export_menu",
        competitor_export="css=competitor_export",
    )


def test_profile_has_competitor_locators():
    assert _competitor_profile().has_competitor_locators() is True
    without = SellerSpriteLocatorProfile(
        panel_open="css=a", ready="css=a", login_required="css=a",
        permission_required="css=a", captcha="css=a", reverse_keywords="css=a",
        asin_input="css=a", submit="css=a", results_ready="css=a",
        export_menu="css=a", export="css=a",
    )
    assert without.has_competitor_locators() is False


def test_profile_from_json_validates_competitor_locators(tmp_path):
    import json

    base = {name: "css=x" for name in (
        "panel_open", "ready", "login_required", "permission_required", "captcha",
        "reverse_keywords", "asin_input", "submit", "results_ready", "export_menu", "export",
    )}
    good = tmp_path / "good.json"
    good.write_text(json.dumps({**base, "competitor_submit": "css=submit"}), encoding="utf-8")
    profile = SellerSpriteLocatorProfile.from_json(good)
    assert profile.competitor_submit == "css=submit"

    bad = tmp_path / "bad.json"
    bad.write_text(json.dumps({**base, "competitor_submit": "not-a-locator"}), encoding="utf-8")
    with pytest.raises(ValueError):
        SellerSpriteLocatorProfile.from_json(bad)


def test_export_competitor_products_drives_locators(tmp_path):
    from agent.tools.sellersprite_browser import PlaywrightSellerSpriteSession

    class FakeLocator:
        def __init__(self, page, name):
            self.page = page
            self.name = name

        def is_visible(self):
            return self.name in self.page.visible

        def click(self, **_kwargs):
            self.page.clicked.append(self.name)

        def fill(self, value, **_kwargs):
            self.page.filled[self.name] = value

    class FakePage:
        def __init__(self):
            self.visible = {
                "competitor_lookup", "competitor_keyword_input", "competitor_submit",
                "competitor_results_ready", "competitor_export_menu", "competitor_export",
            }
            self.clicked = []
            self.filled = {}

        def locator(self, selector):
            return FakeLocator(self, selector.removeprefix("css="))

        def wait_for_timeout(self, _ms):
            return None

    artifact = DownloadedArtifact.from_path(
        _write_csv(tmp_path / "dl.csv", ["Seller", "Price", "Monthly Sales"], [["S", "20", "150"]])
    )

    class FakeObserver:
        def snapshot(self, _path):
            return "snap"

        def wait(self, _path, _snapshot, _timeout, *, cancel_check=None):
            return artifact

    page = FakePage()
    session = PlaywrightSellerSpriteSession(
        profile=_competitor_profile(),
        download_dir=tmp_path,
        page=page,
        download_observer=FakeObserver(),
    )
    result = session.export_competitor_products("bird house")
    assert result is artifact
    assert page.filled["competitor_keyword_input"] == "bird house"
    assert "competitor_submit" in page.clicked
    assert "competitor_export" in page.clicked
