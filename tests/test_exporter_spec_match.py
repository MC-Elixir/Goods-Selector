"""Exporter coverage for supplier match metadata."""
from __future__ import annotations

import json
from types import SimpleNamespace

from reports.exporter import export_excel, export_json, export_markdown


def test_export_json_preserves_supplier_spec_match(tmp_path):
    supplier = SimpleNamespace(
        alibaba_offer_id="1001",
        supplier_name="匹配工厂",
        offer_url="https://detail.1688.com/offer/1001.html",
        base_price_cny=20,
        moq=2,
        monthly_sales=300,
        repeat_buyer_rate=0.3,
        is_factory=True,
        delivery_days=None,
        title_cn="304不锈钢保温杯 700ml 2只装",
        offer_image_url="https://example.com/a.jpg",
        image_similarity=0.91,
        match_quality_score=0.82,
        match_verification_method="heuristic",
        raw_data={
            "source": "alibaba_pifatuan",
            "supplier_quality_score": 0.76,
            "supplier_business_score": 0.93,
            "supplier_candidate_score": 0.84,
            "spec_match": {"score": 0.9, "matched": ["capacity"], "missing": [], "conflicts": []},
            "visual_match": {"score": 0.91, "source": "image_similarity"},
        },
    )
    record = SimpleNamespace(
        product=SimpleNamespace(
            asin="B0TEST",
            marketplace="US",
            title="Water Bottle",
            brand="Generic",
            category="Home & Kitchen",
            price=29.99,
            bsr_rank=1000,
            rating=4.5,
            review_count=100,
            weight_kg=0.4,
            length_cm=7,
            width_cm=7,
            height_cm=25,
            main_image_url=None,
            listing_url=None,
        ),
        profit=None,
        score=None,
        market=SimpleNamespace(
            asin="B0TEST",
            marketplace="US",
            bsr=1000,
            est_monthly_sales=540,
            main_keyword="water bottle",
            search_volume_monthly=4200,
            monthly_purchases=280,
            purchase_rate=6.67,
            raw_data={"asin_detail": {"asin": "B0TEST"}},
        ),
        suppliers=[supplier],
    )

    path = export_json([record], output_path=tmp_path / "candidates.json")
    data = json.loads(path.read_text(encoding="utf-8"))

    exported = data[0]["suppliers"][0]
    assert exported["image_similarity"] == 0.91
    assert exported["match_quality_score"] == 0.82
    assert exported["match_verification_method"] == "heuristic"
    assert exported["supplier_quality_score"] == 0.76
    assert exported["supplier_business_score"] == 0.93
    assert exported["candidate_score"] == 0.84
    assert exported["sourcing_source"] == "alibaba_pifatuan"
    assert exported["raw_data"]["supplier_quality_score"] == 0.76
    assert exported["raw_data"]["supplier_business_score"] == 0.93
    assert exported["raw_data"]["supplier_candidate_score"] == 0.84
    assert exported["raw_data"]["spec_match"]["score"] == 0.9
    assert exported["raw_data"]["visual_match"]["score"] == 0.91
    assert data[0]["market"]["est_monthly_sales"] == 540
    assert data[0]["market"]["main_keyword"] == "water bottle"
    assert data[0]["market"]["raw_data"]["asin_detail"]["asin"] == "B0TEST"


def test_export_excel_includes_visual_and_match_scores(tmp_path):
    supplier = SimpleNamespace(
        alibaba_offer_id="1001",
        supplier_name="匹配工厂",
        offer_url="https://detail.1688.com/offer/1001.html",
        base_price_cny=20,
        moq=2,
        image_similarity=0.91,
        match_quality_score=0.82,
        raw_data={
            "source": "alibaba_pifatuan",
            "supplier_quality_score": 0.76,
            "supplier_business_score": 0.93,
            "supplier_candidate_score": 0.84,
        },
    )
    record = SimpleNamespace(
        product=SimpleNamespace(
            asin="B0TEST",
            title="Water Bottle",
            brand="Generic",
            category="Home & Kitchen",
            price=29.99,
            bsr_rank=1000,
            rating=4.5,
            review_count=100,
        ),
        profit=None,
        score=None,
        suppliers=[supplier],
    )

    path = export_excel([record], output_path=tmp_path / "candidates.xlsx")

    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[2]]
    assert "Top1视觉相似" in headers
    assert "Top1匹配分" in headers
    assert "Top1候选分" in headers
    assert "Top1供应商质量分" in headers
    assert "Top1业务条件分" in headers
    assert "Top1来源" in headers
    assert row[headers.index("Top1来源")] == "alibaba_pifatuan"
    assert row[headers.index("Top1视觉相似")] == 0.91
    assert row[headers.index("Top1匹配分")] == 0.82
    assert row[headers.index("Top1候选分")] == 0.84
    assert row[headers.index("Top1供应商质量分")] == 0.76
    assert row[headers.index("Top1业务条件分")] == 0.93


def test_export_excel_adds_complete_product_supplier_scoring_matrix(tmp_path):
    supplier_one = SimpleNamespace(
        alibaba_offer_id="1001",
        supplier_name="优选源头工厂",
        offer_url="https://detail.1688.com/offer/1001.html",
        title_cn="304不锈钢保温杯 700ml",
        base_price_cny=20,
        moq=2,
        monthly_sales=300,
        repeat_buyer_rate=0.31,
        is_factory=True,
        image_similarity=0.91,
        match_quality_score=0.82,
        match_verification_method="heuristic",
        raw_data={
            "source": "sellersprite_1688",
            "spec_match": {"score": 0.9, "matched": ["capacity"], "missing": [], "conflicts": []},
            "supplier_candidate_score": 0.84,
            "supplier_quality_score": 0.76,
            "supplier_business_score": 0.93,
            "supplier_profit_score": 0.9,
            "supplier_rank_score": 0.86,
            "supplier_profit_margin": 0.35,
            "supplier_net_profit": 10.49,
            "supplier_purchase_cost": 2.8,
            "supplier_shipping_cost": 2.0,
            "supplier_fba_fee": 5.0,
            "supplier_commission": 4.5,
            "supplier_ad_cost": 2.0,
            "supplier_return_loss": 1.0,
            "supplier_exchange_loss": 0.5,
            "supplier_total_cost": 19.5,
        },
    )
    supplier_two = SimpleNamespace(
        alibaba_offer_id="1002",
        supplier_name="待核验供应商",
        offer_url="https://detail.1688.com/offer/1002.html",
        title_cn="保温杯",
        base_price_cny=None,
        moq=None,
        monthly_sales=None,
        repeat_buyer_rate=None,
        is_factory=None,
        image_similarity=None,
        match_quality_score=0.55,
        match_verification_method="heuristic",
        raw_data={
            "source": "sellersprite_1688",
            "spec_match": {"score": 0.6, "matched": [], "missing": ["material"], "conflicts": []},
            "supplier_candidate_score": 0.5,
            "supplier_rank_score": 0.35,
        },
    )
    score = SimpleNamespace(
        total_score=82.5, passed_hard_filter=True, rejection_reasons=[],
        profit_score=0.8, demand_score=0.7, competition_score=0.6,
        supply_score=0.75, logistics_score=0.9, risk_score=0.95,
    )
    record = SimpleNamespace(
        product=SimpleNamespace(
            asin="B0MATRIX", title="Water Bottle", brand="Generic", category="Home",
            price=29.99, bsr_rank=1000, rating=4.5, review_count=100,
            listing_url="https://www.amazon.com/dp/B0MATRIX",
        ),
        profit=None, score=score, suppliers=[supplier_one, supplier_two],
    )

    path = export_excel([record], output_path=tmp_path / "matrix.xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["候选选品", "完整匹配评分表"]
    ws = wb["完整匹配评分表"]
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, (cell.value for cell in row))) for row in ws.iter_rows(min_row=2)]
    assert len(rows) == 2
    assert rows[0]["建议"] == "推荐"
    assert rows[0]["货源来源"] == "sellersprite_1688"
    assert rows[0]["匹配+利润综合分"] == 0.86
    assert rows[0]["预估利润率"] == 0.35
    assert rows[1]["建议"] == "待核验"
    assert rows[1]["1688采购价(CNY)"] is None
    assert rows[1]["预估利润率"] is None


def test_export_markdown_includes_supplier_candidate_scores(tmp_path):
    supplier = SimpleNamespace(
        alibaba_offer_id="1001",
        supplier_name="匹配工厂",
        offer_url="https://detail.1688.com/offer/1001.html",
        base_price_cny=20,
        moq=2,
        monthly_sales=300,
        repeat_buyer_rate=0.3,
        match_quality_score=0.82,
        raw_data={
            "source": "alibaba_pifatuan",
            "supplier_quality_score": 0.76,
            "supplier_candidate_score": 0.84,
        },
    )
    score = SimpleNamespace(
        total_score=82.5,
        passed_hard_filter=True,
        rejection_reasons=[],
        profit_score=0.8,
        demand_score=0.7,
        competition_score=0.6,
        supply_score=0.75,
        logistics_score=0.9,
        risk_score=0.95,
    )
    profit = SimpleNamespace(
        selling_price=29.99,
        purchase_cost=4.0,
        shipping_cost=2.0,
        fba_fee=5.0,
        commission=4.5,
        ad_cost=2.0,
        return_loss=1.0,
        exchange_loss=0.5,
        net_profit=10.99,
        profit_margin=0.3665,
    )
    record = SimpleNamespace(
        product=SimpleNamespace(
            asin="B0TEST",
            title="Water Bottle",
            category="Home & Kitchen",
            price=29.99,
            bsr_rank=1000,
            rating=4.5,
            review_count=100,
            brand="Generic",
            listing_url="https://amazon.example/B0TEST",
            main_image_url=None,
        ),
        profit=profit,
        score=score,
        suppliers=[supplier],
    )

    paths = export_markdown([record], output_dir=tmp_path)
    text = paths[0].read_text(encoding="utf-8")

    assert "候选分" in text
    assert "供应商质量分" in text
    assert "alibaba_pifatuan" in text
    assert "| 0.820 | 0.840 | 0.760 |" in text


def test_insufficient_review_reasons_and_status_reach_all_exports(tmp_path):
    record = SimpleNamespace(
        product=SimpleNamespace(
            asin="B0INSUFFICIENT",
            title="Needs Evidence",
            brand="Generic",
            category="Home & Kitchen",
            price=29.99,
            bsr_rank=1000,
            rating=4.5,
            review_count=100,
            marketplace="US",
        ),
        profit=None,
        score=None,
        market=None,
        suppliers=[],
        rejection_reasons=["missing_purchase_price", "missing_moq"],
    )

    json_path = export_json([record], output_path=tmp_path / "review.json")
    payload = json.loads(json_path.read_text(encoding="utf-8"))[0]
    assert payload["review_status"] == "insufficient_evidence"
    assert payload["rejection_reasons"] == ["missing_purchase_price", "missing_moq"]

    excel_path = export_excel([record], output_path=tmp_path / "review.xlsx")
    import openpyxl
    sheet = openpyxl.load_workbook(excel_path).active
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    assert row[headers.index("审核状态")] == "insufficient_evidence"
    assert row[headers.index("拒绝原因")] == "missing_purchase_price, missing_moq"

    markdown_path = export_markdown([record], output_dir=tmp_path / "markdown")[0]
    markdown = markdown_path.read_text(encoding="utf-8")
    assert "insufficient_evidence" in markdown
    assert "missing_purchase_price, missing_moq" in markdown
