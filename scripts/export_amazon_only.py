"""一次性导出脚本：仅导出已抓取的 Amazon 产品清单（不依赖 1688 匹配）。

适用场景：1688 搜索被 TMD 风控硬拦截、无法完成供应商匹配时，先把 Stage 1
已经抓取并落库的 Amazon BSR 产品导出成 Excel + CSV，供人工查看。

用法：
    python scripts/export_amazon_only.py --category "Garden & Outdoor" --hours 3

说明：
    - 仅读取 SQLite 中已持久化的 products 表，不发起任何网络请求。
    - 通过 category + 最近更新时间窗口筛选本次抓取的产品。
    - 输出到 data/exports/amazon_only_<category>_<timestamp>.{xlsx,csv}。
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime, timedelta
from pathlib import Path

# 允许直接 `python scripts/...` 运行：把仓库根加入 sys.path。
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from sqlalchemy import desc  # noqa: E402

from db.models import Product  # noqa: E402
from db.session import session_scope  # noqa: E402

EXPORT_DIR = ROOT / "data" / "exports"

COLUMNS = [
    ("bsr_rank", "BSR"),
    ("asin", "ASIN"),
    ("title", "标题"),
    ("brand", "品牌"),
    ("price", "售价(USD)"),
    ("rating", "评分"),
    ("review_count", "评论数"),
    ("review_velocity_30d", "30天评论增速"),
    ("category", "类目"),
    ("subcategory", "子类目"),
    ("weight_kg", "重量(kg)"),
    ("length_cm", "长(cm)"),
    ("width_cm", "宽(cm)"),
    ("height_cm", "高(cm)"),
    ("listing_url", "Listing链接"),
]


def _slug(text: str) -> str:
    keep = []
    for ch in text:
        keep.append(ch if ch.isalnum() or ch in "-_" else "_")
    slug = "".join(keep).strip("_")
    return slug or "category"


def _value(row: Product, attr: str):
    value = getattr(row, attr, None)
    if isinstance(value, float):
        return round(value, 4)
    return value


def export(category: str, hours: float) -> dict[str, str]:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    cutoff = datetime.utcnow() - timedelta(hours=hours)

    with session_scope() as s:
        query = s.query(Product).filter(Product.last_updated_at >= cutoff)
        if category:
            query = query.filter(Product.category == category)
        rows = query.order_by(desc(Product.bsr_rank.is_(None)), Product.bsr_rank).all()
        # 转成普通字典，脱离 session 后仍可访问。
        records = [
            {attr: _value(row, attr) for attr, _ in COLUMNS}
            for row in rows
        ]

    if not records:
        logger.warning(f"未找到符合条件的产品（category={category!r}, 最近 {hours} 小时）")
        return {}

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"amazon_only_{_slug(category)}_{stamp}"
    xlsx_path = EXPORT_DIR / f"{base}.xlsx"
    csv_path = EXPORT_DIR / f"{base}.csv"

    # ── Excel ─
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon BSR"
    headers = [label for _, label in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for rec in records:
        ws.append([rec[attr] for attr, _ in COLUMNS])
    widths = [10, 14, 60, 16, 12, 8, 10, 12, 18, 18, 10, 8, 8, 8, 50]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)

    # ── CSV（UTF-8 BOM，便于 Excel 直接打开中文）──
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for rec in records:
            writer.writerow([rec[attr] for attr, _ in COLUMNS])

    logger.info(f"导出 {len(records)} 个 Amazon 产品")
    logger.info(f"  Excel: {xlsx_path}")
    logger.info(f"  CSV  : {csv_path}")
    return {"xlsx": str(xlsx_path), "csv": str(csv_path), "count": len(records)}  # type: ignore[dict-item]


def main() -> None:
    parser = argparse.ArgumentParser(description="导出已抓取的 Amazon 产品清单（不含 1688 匹配）")
    parser.add_argument("--category", default="Garden & Outdoor", help="Amazon 类目，用于筛选")
    parser.add_argument("--hours", type=float, default=3.0, help="只导出最近 N 小时内更新的产品")
    args = parser.parse_args()

    result = export(args.category, args.hours)
    if not result:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
